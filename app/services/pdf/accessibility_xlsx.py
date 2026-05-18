"""Excel export for enriched accessibility reports."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def accessibility_report_to_xlsx_bytes(summary: dict[str, Any], rows: list[dict[str, str]]) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.cell(row=1, column=1, value="Field").font = Font(bold=True)
    ws_summary.cell(row=1, column=2, value="Value").font = Font(bold=True)
    r = 2
    for key, value in summary.items():
        ws_summary.cell(row=r, column=1, value=str(key))
        ws_summary.cell(row=r, column=2, value=value if isinstance(value, (int, float, bool)) else str(value))
        r += 1
    if r == 2:
        ws_summary.cell(row=2, column=1, value="(no summary data)")

    ws_detail = wb.create_sheet("Detailed findings")
    headers = ["Section", "Rule", "Status", "Description", "Pages", "Notes"]
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws_detail.append(
            [
                row.get("section", ""),
                row.get("rule", ""),
                row.get("status", ""),
                row.get("description", ""),
                row.get("pages", ""),
                row.get("notes", ""),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def stamp_department_metadata(xlsx_path: str | Path, department_name: str) -> None:
    """Add or update a Metadata sheet on an existing workbook (e.g. Adobe tagging report)."""
    path = Path(xlsx_path)
    if not path.is_file():
        return
    wb = load_workbook(path)
    if "Metadata" in wb.sheetnames:
        ws = wb["Metadata"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Metadata", 0)
    ws.cell(row=1, column=1, value="Field").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Value").font = Font(bold=True)
    ws.cell(row=2, column=1, value="Department")
    ws.cell(row=2, column=2, value=department_name or "Default")
    wb.save(path)
